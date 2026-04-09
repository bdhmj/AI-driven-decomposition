import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GANTT Chart"

csv_data = [
    # ── Infrastructure (DevOps) ──────────────────────────────────────────
    ("Infrastructure (DevOps)", "Создание и настройка репозиториев",          "2025-06-02", "2025-06-03"),
    ("Infrastructure (DevOps)", "Настройка CI/CD pipelines",                 "2025-06-05", "2025-06-07"),
    ("Infrastructure (DevOps)", "Настройка PostgreSQL",                      "2025-06-08", "2025-06-10"),
    ("Infrastructure (DevOps)", "Настройка Redis",                           "2025-06-11", "2025-06-13"),
    ("Infrastructure (DevOps)", "Настройка HashiCorp Vault",                 "2025-06-14", "2025-06-16"),
    ("Infrastructure (DevOps)", "Настройка Nginx и доменов",                 "2025-06-17", "2025-07-21"),
    ("Infrastructure (DevOps)", "Настройка централизованного логирования",   "2025-07-23", "2025-07-26"),
    ("Infrastructure (DevOps)", "Настройка мониторинга и алертинга",         "2025-07-27", "2025-07-30"),

    # ── Smart contracts ──────────────────────────────────────────────────
    ("Smart contracts", "Разработка Escrow Implementation",                  "2025-06-02", "2025-06-05"),
    ("Smart contracts", "Разработка EscrowFactory",                          "2025-06-06", "2025-06-07"),
    ("Smart contracts", "Написание unit и integration тестов",               "2025-06-08", "2025-06-11"),
    ("Smart contracts", "Деплой в testnet (Sepolia)",                        "2025-06-12", "2025-06-12"),
    ("Smart contracts", "Аудит безопасности",                                "2025-06-13", "2025-07-30"),

    # ── Backend #3 ───────────────────────────────────────────────────────
    ("Backend #3", "Настройка инфраструктуры",                               "2025-06-02", "2025-06-05"),
    ("Backend #3", "Разработка Shared Layer",                                "2025-06-06", "2025-06-09"),
    ("Backend #3", "Разработка Blockchain Service",                          "2025-06-10", "2025-06-13"),
    ("Backend #3", "Trading Module (сущности)",                              "2025-06-14", "2025-06-15"),
    ("Backend #3", "Trading Module (order API)",                             "2025-06-16", "2025-06-21"),
    ("Backend #3", "Trading Module (deal API)",                              "2025-06-22", "2025-06-27"),
    ("Backend #3", "Trading Module (Lifecycle сделки)",                      "2025-06-28", "2025-07-03"),
    ("Backend #3", "Разработка Encryption Service",                          "2025-07-04", "2025-07-05"),
    ("Backend #3", "Payment Module (сущность PaymentDetails)",               "2025-07-06", "2025-07-06"),
    ("Backend #3", "Payment Module (API)",                                   "2025-07-07", "2025-07-08"),
    ("Backend #3", "Trading Module (реализация «ямы»)",                      "2025-07-09", "2025-07-14"),
    ("Backend #3", "Миграции БД (создание миграций)",                        "2025-07-15", "2025-07-20"),
    ("Backend #3", "Миграции БД (Настройка Swagger/OpenAPI)",                "2025-07-21", "2025-07-22"),
    ("Backend #3", "Unit-тестирование Backend",                              "2025-07-23", "2025-07-26"),

    # ── Backend #4 ───────────────────────────────────────────────────────
    ("Backend #4", "User Module (Создание сущностей)",                       "2025-06-03", "2025-06-05"),
    ("Backend #4", "User Module (API)",                                      "2025-06-06", "2025-06-10"),
    ("Backend #4", "User Module (уровни доступа)",                           "2025-06-11", "2025-06-14"),
    ("Backend #4", "Чат интеграция",                                         "2025-06-15", "2025-06-19"),
    ("Backend #4", "Dispute Module (сущности)",                              "2025-06-20", "2025-06-21"),
    ("Backend #4", "Dispute Module (API)",                                   "2025-06-22", "2025-06-25"),
    ("Backend #4", "Dispute Module (admin API)",                             "2025-06-26", "2025-06-29"),
    ("Backend #4", "Referral Module (API реферальной ссылки)",               "2025-06-30", "2025-06-30"),
    ("Backend #4", "Referral Module (API кошелька комиссий)",                "2025-07-01", "2025-07-01"),
    ("Backend #4", "Referral Module (Привязка при регистрации)",             "2025-07-02", "2025-07-02"),
    ("Backend #4", "Referral Module (Интеграция со сделками)",               "2025-07-03", "2025-07-04"),
    ("Backend #4", "Referral Module (сущность ReferralReward)",              "2025-07-05", "2025-07-05"),
    ("Backend #4", "Referral Module (API реферальной статистики)",           "2025-07-06", "2025-07-07"),
    ("Backend #4", "Referral Module (Обработка событий выплат)",             "2025-07-08", "2025-07-09"),
    ("Backend #4", "KYC-интеграция",                                         "2025-07-10", "2025-07-13"),
    ("Backend #4", "Rating Module (сущности)",                               "2025-07-14", "2025-07-15"),
    ("Backend #4", "Rating Module (сервис рейтинга)",                        "2025-07-16", "2025-07-17"),
    ("Backend #4", "Telegram (уведомление о статусе сделки)",                "2025-07-18", "2025-07-19"),
    ("Backend #4", "Telegram (OTC сделки в канал)",                          "2025-07-19", "2025-07-19"),
    ("Backend #4", "Analytics Module (подключение аналитики)",               "2025-07-20", "2025-07-20"),

    # ── Frontend ─────────────────────────────────────────────────────────
    ("Frontend", "Инициализация",                                            "2025-06-02", "2025-06-07"),
    ("Frontend", "Страница авторизации",                                     "2025-06-08", "2025-06-09"),
    ("Frontend", "Страница профиля",                                         "2025-06-10", "2025-06-13"),
    ("Frontend", "Страница стакана (Order Book)",                            "2025-06-14", "2025-06-20"),
    ("Frontend", "Страница Создания Ордера",                                 "2025-06-21", "2025-06-24"),
    ("Frontend", "Страница сделки + чат",                                    "2025-06-25", "2025-07-08"),
    ("Frontend", "Referral Module (раздел в ЛК)",                            "2025-07-09", "2025-07-10"),
    ("Frontend", "Referral Module (Таблица сделок)",                         "2025-07-11", "2025-07-12"),
    ("Frontend", "Referral Module (Интеграция API)",                         "2025-07-13", "2025-07-14"),
    ("Frontend", "Страница Истории Сделок",                                  "2025-07-15", "2025-07-15"),
    ("Frontend", "KYC интеграция",                                           "2025-07-16", "2025-07-17"),
    ("Frontend", "Страница «Яма» (Pit)",                                     "2025-07-18", "2025-07-19"),
    ("Frontend", "Интеграция WebSocket",                                     "2025-07-20", "2025-07-21"),
    ("Frontend", "Админ-панель (Layout + Dashboard)",                        "2025-07-22", "2025-07-29"),
    ("Frontend", "Доработка интерфейсов под мобильные",                      "2025-07-30", "2025-08-04"),

    # ── QA ───────────────────────────────────────────────────────────────
    ("QA", "Интеграционное тестирование",                                    "2025-06-02", "2025-06-25"),
    ("QA", "Тестирование безопасности",                                      "2025-06-26", "2025-07-10"),
    ("QA", "Багфикси и регрессия",                                           "2025-07-11", "2025-08-10"),

    # ── Analyst ──────────────────────────────────────────────────────────
    ("Analyst", "Сопровождение",                                             "2025-06-02", "2025-08-10"),
]

# ══════════════════════════════════════════════════════════════════════════════

def next_workday(dt):
    while dt.weekday() >= 5:
        dt += timedelta(days=1)
    return dt

def add_workdays(start_dt, num_workdays):
    cur = next_workday(start_dt)
    counted = 1
    while counted < num_workdays:
        cur += timedelta(days=1)
        if cur.weekday() < 5:
            counted += 1
    return cur

# ── Cascade (v6 logic) ─────────────────────────────────────────────────────
tasks = []
phase_prev = {}

for phase, title, orig_start_str, orig_end_str in csv_data:
    os_dt = datetime.strptime(orig_start_str, "%Y-%m-%d")
    oe_dt = datetime.strptime(orig_end_str, "%Y-%m-%d")
    duration = (oe_dt - os_dt).days + 1

    if phase in phase_prev:
        prev_orig_end, prev_new_end = phase_prev[phase]
        gap = (os_dt - prev_orig_end).days
        if gap <= 1:
            new_start = next_workday(prev_new_end + timedelta(days=gap))
        else:
            accumulated = (prev_new_end - prev_orig_end).days
            new_start = next_workday(os_dt + timedelta(days=accumulated))
    else:
        new_start = next_workday(os_dt)

    new_end = add_workdays(new_start, duration)
    tasks.append((phase, title, new_start, new_end, duration))
    phase_prev[phase] = (oe_dt, new_end)

# ══════════════════════════════════════════════════════════════════════════════
phase_to_role = {
    "Infrastructure (DevOps)": "DevOps", "Smart contracts": "Smart Contract",
    "Backend #3": "Backend #3", "Backend #4": "Backend #4",
    "Frontend": "Frontend", "QA": "QA", "Analyst": "Analyst",
}

project_start = min(t[2] for t in tasks)
while project_start.weekday() != 0:
    project_start -= timedelta(days=1)
project_end = max(t[3] for t in tasks)
while project_end.weekday() != 4:
    project_end += timedelta(days=1)

all_days = []
d = project_start
while d <= project_end:
    all_days.append(d)
    d += timedelta(days=1)
num_days = len(all_days)
DATA_COL_START = 7

phase_colors = {
    "Infrastructure (DevOps)": {"header": "1F4E79", "fill": "D6E4F0", "bar": "5B9BD5"},
    "Smart contracts":         {"header": "7B2D26", "fill": "F2DCDB", "bar": "C0504D"},
    "Backend #3":              {"header": "4F6228", "fill": "EBF1DE", "bar": "9BBB59"},
    "Backend #4":              {"header": "31859C", "fill": "DAEEF3", "bar": "4BACC6"},
    "Frontend":                {"header": "E36C09", "fill": "FDE9D9", "bar": "F79646"},
    "QA":                      {"header": "60497A", "fill": "E4DFEC", "bar": "8064A2"},
    "Analyst":                 {"header": "4A452A", "fill": "F2F2E6", "bar": "948A54"},
}

thin_border = Border(
    left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'),
)
week_sep_border = Border(
    left=Side(style='thin', color='BFBFBF'), right=Side(style='medium', color='808080'),
    top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'),
)

header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
day_num_font = Font(name='Calibri', size=7, color='44546A')
task_font = Font(name='Calibri', size=9)
phase_header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

weekend_header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
date_label_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
center_align = Alignment(horizontal='center', vertical='center')
left_wrap = Alignment(vertical='center', wrap_text=True)

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 42
ws.column_dimensions['D'].width = 11
ws.column_dimensions['E'].width = 6
ws.column_dimensions['F'].width = 11
for i in range(num_days):
    cl = get_column_letter(DATA_COL_START + i)
    ws.column_dimensions[cl].width = 2.5 if all_days[i].weekday() >= 5 else 3.8

# ROW 1: Month headers
ws.row_dimensions[1].height = 20
for c in range(1, DATA_COL_START):
    cell = ws.cell(row=1, column=c); cell.fill = header_fill; cell.border = thin_border

month_names_ru = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
    7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
month_spans = []; cur_month = None; span_start = None
for i, day in enumerate(all_days):
    mk = (day.year, day.month); col = DATA_COL_START + i
    if mk != cur_month:
        if cur_month is not None: month_spans.append((span_start, col-1, cur_month))
        cur_month = mk; span_start = col
    if i == len(all_days)-1: month_spans.append((span_start, col, cur_month))
for sc, ec, (year, month) in month_spans:
    if ec > sc: ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
    cell = ws.cell(row=1, column=sc, value=f"{month_names_ru[month]} {year}")
    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
    for cc in range(sc, ec+1):
        ws.cell(row=1, column=cc).fill = header_fill; ws.cell(row=1, column=cc).border = thin_border

# ROW 2: Headers + day numbers
ws.row_dimensions[2].height = 18
for ci, h in enumerate(["Фаза","Роль","Задача","Старт","Дней","Конец"], 1):
    cell = ws.cell(row=2, column=ci, value=h)
    cell.font = header_font; cell.fill = header_fill; cell.alignment = center_align; cell.border = thin_border
for i, day in enumerate(all_days):
    col = DATA_COL_START + i; is_wknd = day.weekday() >= 5
    cell = ws.cell(row=2, column=col, value=day.day)
    cell.font = day_num_font; cell.fill = weekend_header_fill if is_wknd else date_label_fill
    cell.alignment = center_align; cell.border = week_sep_border if day.weekday()==6 else thin_border

# ROW 3: Day-of-week
ws.row_dimensions[3].height = 14
day_abbr = ["Пн","Вт","Ср","Чт","Пт","Сб","Вс"]
for ci in range(1, DATA_COL_START):
    cell = ws.cell(row=3, column=ci); cell.fill = date_label_fill; cell.border = thin_border
for i, day in enumerate(all_days):
    col = DATA_COL_START + i; is_wknd = day.weekday() >= 5
    cell = ws.cell(row=3, column=col, value=day_abbr[day.weekday()])
    cell.font = Font(name='Calibri', size=6, color='999999' if is_wknd else '808080', bold=is_wknd)
    cell.fill = weekend_header_fill if is_wknd else date_label_fill
    cell.alignment = center_align; cell.border = week_sep_border if day.weekday()==6 else thin_border

# DATA ROWS
current_row = 4
current_phase = None
for phase, title, start_dt, end_dt, duration in tasks:
    colors = phase_colors.get(phase, phase_colors["Analyst"])
    role = phase_to_role.get(phase, phase)
    if phase != current_phase:
        current_phase = phase
        pf = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type='solid')
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1, value=phase)
        cell.font = phase_header_font; cell.fill = pf
        cell.alignment = Alignment(horizontal='left', vertical='center'); cell.border = thin_border
        for c in range(2, DATA_COL_START):
            ws.cell(row=current_row, column=c).fill = pf; ws.cell(row=current_row, column=c).border = thin_border
        for i, day in enumerate(all_days):
            c = ws.cell(row=current_row, column=DATA_COL_START+i); c.fill = pf
            c.border = week_sep_border if day.weekday()==6 else thin_border
        ws.row_dimensions[current_row].height = 22; current_row += 1

    row_fill = PatternFill(start_color=colors["fill"], end_color=colors["fill"], fill_type='solid')
    bar_fill = PatternFill(start_color=colors["bar"], end_color=colors["bar"], fill_type='solid')
    rb,gb,bb = int(colors["fill"][:2],16), int(colors["fill"][2:4],16), int(colors["fill"][4:6],16)
    wknd_row = PatternFill(
        start_color=f"{max(0,rb-25):02X}{max(0,gb-25):02X}{max(0,bb-25):02X}",
        end_color=f"{max(0,rb-25):02X}{max(0,gb-25):02X}{max(0,bb-25):02X}", fill_type='solid')

    row_data = [phase, role, title, start_dt.strftime("%d.%m.%y"), duration, end_dt.strftime("%d.%m.%y")]
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=ci, value=val)
        cell.font = task_font; cell.fill = row_fill; cell.border = thin_border
        cell.alignment = center_align if ci in (4,5,6) else left_wrap

    for i, day in enumerate(all_days):
        col = DATA_COL_START + i; cell = ws.cell(row=current_row, column=col)
        is_wknd = day.weekday() >= 5; in_range = start_dt <= day <= end_dt
        cell.border = week_sep_border if day.weekday()==6 else thin_border
        if in_range and day.weekday() < 5: cell.fill = bar_fill
        elif is_wknd: cell.fill = wknd_row
        else: cell.fill = row_fill

    ws.row_dimensions[current_row].height = 20; current_row += 1

ws.freeze_panes = 'G4'
output_path = r"C:\Users\gregory\Desktop\P2P_Roadmap_Gantt_v11b.xlsx"
wb.save(output_path)
print(f"Saved to {output_path}")
print(f"Range: {project_start.strftime('%d.%m.%y')} - {project_end.strftime('%d.%m.%y')} ({num_days} cal days)")
print()
print("=== PHASE END DATES ===")
pe = {}
for phase, title, s, e, dur in tasks:
    if phase not in pe or e > pe[phase]: pe[phase] = e
for phase, end in sorted(pe.items(), key=lambda x: x[1]):
    print(f"  {phase:30s} {end.strftime('%d.%m.%y')}")
