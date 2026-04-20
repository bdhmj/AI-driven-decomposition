"""Build project estimation .xlsx report from decomposition JSON.

Usage:
    python scripts/build_xlsx.py input/decomposition.json output/Оценка_проекта.xlsx \
        --params output/estimation_params.json [--name "Project Name"]

In full mode (--params) generates 6 sheets:
  "Для клиента", "Sales (Итоги оценки)", "Оценка",
  "PM (Расчеты)", "Для Битрикса", "GANTT Chart"

PM (Расчеты) is the single source of truth — its rates and
coefficient cells (C21..C27) feed every other sheet via formulas.
A PM who tweaks values in PM (Расчеты) will see all downstream
totals, weeks, client-costs and Gantt-day counts recalculate live.

In simple mode (no --params) falls back to 3-sheet days-only output
(backward compatible with old pipeline).
"""

import io
import json
import math
import os
import sys
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════════════
# Constants — sheet names and PM cell layout
# ═══════════════════════════════════════════════════════════════════════

PM_SHEET = "PM (Расчеты)"
SALES_SHEET = "Sales (Итоги оценки)"
ESTIMATION_SHEET = "Оценка"
CLIENT_SHEET = "Для клиента"
BITRIX_SHEET = "Для Битрикса"
GANTT_SHEET = "GANTT Chart"

PM_REF = f"'{PM_SHEET}'"
EST_REF = f"'{ESTIMATION_SHEET}'"
SALES_REF = f"'{SALES_SHEET}'"
CLI_REF = f"'{CLIENT_SHEET}'"

# PM sheet cell layout (FIXED — formulas everywhere lock onto these)
PM_QA_ROW = 5          # Manual QA (auto via qa_pct)
PM_MGR_ROW = 6         # Project manager (auto via pm_pct)
PM_SPEC_START = 7      # Real specialists start here
PM_SPEC_MAX = 12       # Rows 7..18
PM_SPEC_END = PM_SPEC_START + PM_SPEC_MAX - 1  # 18

# Coefficient input cells (column C, hardcoded rows)
PM_COEF = {
    "debug_pct": 21,
    "code_review_hours": 22,
    "communication_hours": 23,
    "qa_pct": 24,
    "risk_buffer_pct": 25,
    "devops_pct": 26,
    "pm_pct": 27,
}

# Оценка sheet layout
EST_SUMMARY_START = 5      # Per-spec rollup begins here
EST_SUMMARY_END = 10       # Ends (6 rows, 12 specs fit in 2 blocks: A-D and E-H)
EST_TASK_HEADER_ROW = 11
EST_TASK_START = 12
EST_TASK_END = 1000        # SUMPRODUCT range upper bound (generous)

# Sales sheet layout
SALES_TOTAL_ROW = 5        # Project totals row
SALES_QA_ROW = 6           # Manual QA
SALES_MGR_ROW = 7          # PM
SALES_SPEC_START = 8       # Real specialists (up to 12, rows 8..19)
SALES_SPEC_END = SALES_SPEC_START + PM_SPEC_MAX - 1  # 19


# ═══════════════════════════════════════════════════════════════════════
# Helpers: Excel formula strings
# ═══════════════════════════════════════════════════════════════════════

def k_expr() -> str:
    """Excel formula string computing K from PM (Расчеты) coefficient cells.

    Mirrors calc_K() in pure Excel:
        K = 1 + code_review/8 + comms/40 + debug/100 + risk/100 + devops/100

    Wrapped in outer parentheses so it's safe to multiply/divide with
    surrounding operators without operator-precedence surprises.
    """
    return (
        f"({PM_REF}!$C${PM_COEF['code_review_hours']}/8"
        f"+{PM_REF}!$C${PM_COEF['communication_hours']}/40"
        f"+{PM_REF}!$C${PM_COEF['debug_pct']}/100"
        f"+{PM_REF}!$C${PM_COEF['risk_buffer_pct']}/100"
        f"+{PM_REF}!$C${PM_COEF['devops_pct']}/100"
        f"+1)"
    )


def calc_K(coeffs: dict) -> float:
    """Python mirror of k_expr(), used where a numeric K is needed (Gantt)."""
    return (
        1
        + coeffs.get("code_review_hours", 0) / 8
        + coeffs.get("communication_hours", 0) / 40
        + coeffs.get("debug_pct", 0) / 100
        + coeffs.get("risk_buffer_pct", 0) / 100
        + coeffs.get("devops_pct", 0) / 100
    )


def half_up(x: float) -> int:
    """Round half away from zero (like Excel ROUND), not banker's.
    Keeps Gantt days in lockstep with =ROUND() formulas elsewhere.
    """
    return math.floor(x + 0.5) if x >= 0 else -math.floor(-x + 0.5)


# ═══════════════════════════════════════════════════════════════════════
# Shared styles
# ═══════════════════════════════════════════════════════════════════════

FILL_ORANGE = PatternFill(start_color="FFA301", end_color="FFA301", fill_type="solid")
FILL_GRAY = PatternFill(start_color="F4F4F4", end_color="F4F4F4", fill_type="solid")
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_GREEN = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # editable
FILL_BLUE_HEADER = PatternFill(start_color="4A86E8", end_color="4A86E8", fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # totals

BORDER_THIN = Border(
    left=Side("thin"), right=Side("thin"),
    top=Side("thin"), bottom=Side("thin"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

FONT_TITLE = Font(name="Arial", size=14, bold=True)
FONT_HEADER = Font(name="Arial", size=11, bold=True)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
FONT_WHITE_BOLD = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_SECTION = Font(name="Arial", size=11, bold=True)


def _apply_border_rect(ws, r1, c1, r2, c2, border=BORDER_THIN):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = border


# ═══════════════════════════════════════════════════════════════════════
# Build entry point
# ═══════════════════════════════════════════════════════════════════════

def build_report_xlsx(
    project_name: str,
    modules: list[dict],
    K: float = 1.0,
    params: dict | None = None,
) -> io.BytesIO:
    """Build project estimation xlsx.

    - With `params`: 6-sheet formula-driven output (PM sheet is source of truth).
    - Without `params`: legacy 3-sheet days-only output.
    """
    wb = Workbook()

    if params is None:
        # Legacy simple mode — keep backward compatibility
        _build_simple_mode(wb, project_name, modules, K)
    else:
        _build_full_mode(wb, project_name, modules, params)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════
# FULL MODE (6 sheets with formulas)
# ═══════════════════════════════════════════════════════════════════════

def _collect_roster(modules: list[dict], params: dict) -> list[dict]:
    """Build the ordered specialist roster for the workbook.

    Priority order:
    1. params["specialists"] if present — honors user's Step 2 confirmation
    2. Otherwise, specialists in decomposition order + any extra rates keys

    PM and Manual QA are NOT real specialists — they live in dedicated rows
    on PM/Sales sheets and are filtered out here.

    Returns list of dicts: {name, index (1..N), pm_row, est_sum_row, est_sum_col}
    est_sum_row/col point to the per-spec summary cell in Оценка (12 spec max,
    indices 1..6 go to A-D block rows 5-10, indices 7..12 go to E-H block rows 5-10).
    """
    def _is_auto(name: str) -> bool:
        return name.strip().lower() in ("pm", "project manager", "manual qa", "qa (auto)")

    ordered: list[str] = []
    seen: set[str] = set()

    # 1. Explicit roster from flow Step 2
    for name in params.get("specialists", []):
        if not _is_auto(name) and name not in seen:
            ordered.append(name)
            seen.add(name)

    # 2. Whatever appeared in decomposition (preserves module traversal order)
    for m in modules:
        for t in m.get("tasks", []):
            name = t["specialist"]
            if not _is_auto(name) and name not in seen:
                ordered.append(name)
                seen.add(name)

    # 3. Anything still in rates not yet added (e.g. DevOps with no tasks yet)
    for name in params.get("rates", {}):
        if not _is_auto(name) and name not in seen:
            ordered.append(name)
            seen.add(name)

    if len(ordered) > PM_SPEC_MAX:
        raise ValueError(
            f"Слишком много специалистов ({len(ordered)}). "
            f"Максимум {PM_SPEC_MAX}. Сократите ростер в Step 2 флоу."
        )

    roster = []
    for i, name in enumerate(ordered):
        index = i + 1                       # 1-based
        pm_row = PM_SPEC_START + i          # 7, 8, ...
        # Estimation summary position:
        # indices 1..6 → rows 5..10, cols A-D
        # indices 7..12 → rows 5..10, cols E-H
        if index <= 6:
            est_sum_row = EST_SUMMARY_START + (index - 1)
            est_sum_col_block = "A"   # anchor column of the block
        else:
            est_sum_row = EST_SUMMARY_START + (index - 7)
            est_sum_col_block = "E"
        roster.append({
            "name": name,
            "index": index,
            "pm_row": pm_row,
            "est_sum_row": est_sum_row,
            "est_sum_block": est_sum_col_block,
        })
    return roster


def _build_full_mode(wb: Workbook, project_name: str, modules: list[dict], params: dict):
    roster = _collect_roster(modules, params)
    rates = params.get("rates", {})
    coefficients = params.get("coefficients", {})
    margin_pct = params.get("margin_pct", 0)  # markup %
    currency = params.get("currency", "$")

    # K for python-side Gantt calc (matches k_expr() numerically)
    K = calc_K(coefficients)

    # Track per-task mapping: module-order task_index → Оценка row
    # Built while writing Оценка, used later by Client & Bitrix.
    task_to_est_row: list[int] = []

    # Sheet order (left→right in Excel tab bar):
    #   Для клиента → Sales → Оценка → PM → Для Битрикса → GANTT
    # Openpyxl adds sheets in creation order; `wb.active` is the first.
    _build_client_sheet(wb, project_name, modules, roster, rates, margin_pct, currency, task_to_est_row)
    _build_sales_sheet(wb, project_name, roster, rates, margin_pct, currency)
    _build_estimation_sheet(wb, modules, roster, coefficients, task_to_est_row)
    _build_pm_sheet(wb, roster, rates, coefficients, currency)
    _build_bitrix_sheet(wb, modules, task_to_est_row)
    _build_gantt_sheet(wb, modules, K, roster)


# ═══════════════════════════════════════════════════════════════════════
# Sheet: PM (Расчеты) — single source of truth
# ═══════════════════════════════════════════════════════════════════════

def _build_pm_sheet(wb, roster, rates, coefficients, currency):
    ws = wb.create_sheet(PM_SHEET)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 32
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 16

    # Instructions
    ws.cell(row=1, column=2, value="🛠 Листа PM (Расчеты)").font = FONT_TITLE
    ws.cell(row=2, column=2, value=(
        "Зелёные ячейки редактируются (ставки и коэффициенты). "
        "Всё остальное — формулы, пересчитываются автоматически."
    )).font = FONT_NORMAL
    ws.row_dimensions[2].height = 24

    # Headers (row 4)
    hdr = {
        2: "Проектная команда / тип работ",
        3: "Индекс",
        4: f"Часовая ставка, {currency}",
        5: f"Ставка клиенту, {currency}",
        7: "Проектная команда",
        8: "Индекс",
        9: "Продолжительность, недель",
        10: "Себестоимость",
    }
    for col, val in hdr.items():
        c = ws.cell(row=4, column=col, value=val)
        c.font = FONT_WHITE_BOLD
        c.fill = FILL_BLUE_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
    ws.row_dimensions[4].height = 32

    # Row 5: Manual QA (auto) — % of NON-QA specialist weeks.
    # If QA already exists in roster, subtract its own weeks to avoid double-count.
    qa_range_excl = (
        f"(SUM(I{PM_SPEC_START}:I{PM_SPEC_END})"
        f"-SUMIF(B{PM_SPEC_START}:B{PM_SPEC_END},\"QA\",I{PM_SPEC_START}:I{PM_SPEC_END})"
        f"-SUMIF(B{PM_SPEC_START}:B{PM_SPEC_END},\"Manual QA\",I{PM_SPEC_START}:I{PM_SPEC_END}))"
    )
    _pm_write_auto_row(ws, PM_QA_ROW, "Manual QA",
                       rate=rates.get("QA", rates.get("Manual QA", 15)),
                       weeks_formula=f"={qa_range_excl}*$C${PM_COEF['qa_pct']}/100",
                       cost_formula=f"=I{PM_QA_ROW}*5*8*D{PM_QA_ROW}")

    # Row 6: Project Manager (auto)
    _pm_write_auto_row(ws, PM_MGR_ROW, "Project manager",
                       rate=rates.get("PM", 25),
                       weeks_formula=f"=MAX(I{PM_SPEC_START}:I{PM_SPEC_END})*$C${PM_COEF['pm_pct']}/100",
                       cost_formula=f"=I{PM_MGR_ROW}*5*8*D{PM_MGR_ROW}")

    # Rows 7..: real specialists
    for s in roster:
        r = s["pm_row"]
        # B: name (editable-looking but stable)
        cell_b = ws.cell(row=r, column=2, value=s["name"])
        cell_b.font = FONT_NORMAL
        cell_b.alignment = ALIGN_LEFT
        cell_b.border = BORDER_THIN
        # C: index (hardcoded to preserve Оценка SUMPRODUCT contract)
        cell_c = ws.cell(row=r, column=3, value=s["index"])
        cell_c.font = FONT_NORMAL
        cell_c.alignment = ALIGN_CENTER
        cell_c.border = BORDER_THIN
        # D: hourly rate (editable green)
        rate_val = rates.get(s["name"], 20)
        cell_d = ws.cell(row=r, column=4, value=rate_val)
        cell_d.font = FONT_NORMAL
        cell_d.fill = FILL_GREEN
        cell_d.alignment = ALIGN_RIGHT
        cell_d.border = BORDER_THIN
        cell_d.number_format = "0.00"
        # E: client rate (mirror from Sales after markup)
        cell_e = ws.cell(row=r, column=5, value=f"={SALES_REF}!K{SALES_SPEC_START + s['index'] - 1}")
        cell_e.font = FONT_NORMAL
        cell_e.alignment = ALIGN_RIGHT
        cell_e.border = BORDER_THIN
        cell_e.number_format = "0.00"
        # G: name mirror
        cell_g = ws.cell(row=r, column=7, value=f"=B{r}")
        cell_g.font = FONT_NORMAL
        cell_g.alignment = ALIGN_LEFT
        cell_g.border = BORDER_THIN
        # H: index mirror
        cell_h = ws.cell(row=r, column=8, value=s["index"])
        cell_h.font = FONT_NORMAL
        cell_h.alignment = ALIGN_CENTER
        cell_h.border = BORDER_THIN
        # I: weeks from Оценка summary cell (simple formula: avg days × K / 5)
        # For index 1..6: B{est_row}=min sum, C{est_row}=max sum
        # For index 7..12: F{est_row}=min sum, G{est_row}=max sum
        if s["index"] <= 6:
            min_ref = f"{EST_REF}!B{s['est_sum_row']}"
            max_ref = f"{EST_REF}!C{s['est_sum_row']}"
        else:
            min_ref = f"{EST_REF}!F{s['est_sum_row']}"
            max_ref = f"{EST_REF}!G{s['est_sum_row']}"
        weeks_formula = f"=(({min_ref}+{max_ref})/2)*{k_expr()}/5"
        cell_i = ws.cell(row=r, column=9, value=weeks_formula)
        cell_i.font = FONT_NORMAL
        cell_i.alignment = ALIGN_RIGHT
        cell_i.border = BORDER_THIN
        cell_i.number_format = "0.00"
        # J: cost = I * 5 * 8 * D
        cell_j = ws.cell(row=r, column=10, value=f"=I{r}*5*8*D{r}")
        cell_j.font = FONT_NORMAL
        cell_j.alignment = ALIGN_RIGHT
        cell_j.border = BORDER_THIN
        cell_j.number_format = f"#,##0 \"{currency}\""

    # Fill empty slots up to PM_SPEC_END with borders only
    for r in range(PM_SPEC_START + len(roster), PM_SPEC_END + 1):
        for col in [2, 3, 4, 5, 7, 8, 9, 10]:
            ws.cell(row=r, column=col).border = BORDER_THIN
        ws.cell(row=r, column=4).fill = FILL_GREEN

    # Coefficients block — row 20 header, rows 21-27 inputs
    ws.cell(row=20, column=2, value="⚙️ Коэффициенты проекта").font = FONT_TITLE
    ws.row_dimensions[20].height = 22

    coef_labels = [
        ("debug_pct",          "Проверка и отладка задач (%)", "%"),
        ("code_review_hours",  "Код ревью (часов/день на разработчика)", "h"),
        ("communication_hours","Коммуникации (часов/неделю)", "h"),
        ("qa_pct",             "Тестировщик (% от общего)", "%"),
        ("risk_buffer_pct",    "Буфер на риски (%)", "%"),
        ("devops_pct",         "DevOps (%) доп.", "%"),
        ("pm_pct",             "Менеджер (% от максимальной длительности спеца)", "%"),
    ]
    for key, label, unit in coef_labels:
        r = PM_COEF[key]
        ws.cell(row=r, column=2, value=label).font = FONT_NORMAL
        ws.cell(row=r, column=2).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2).border = BORDER_THIN
        cell_c = ws.cell(row=r, column=3, value=coefficients.get(key, 0))
        cell_c.font = FONT_BOLD
        cell_c.fill = FILL_GREEN
        cell_c.alignment = ALIGN_CENTER
        cell_c.border = BORDER_THIN
        cell_c.number_format = "0.###"

    # D-column formulas: weeks attributable to each coefficient (informative)
    # Template: (total_spec_weeks / K) × (coef / divisor)
    total_spec_weeks = f"SUM(I{PM_SPEC_START}:I{PM_SPEC_END})"
    base_no_K = f"({total_spec_weeks}/{k_expr()})"
    d_formulas = {
        "debug_pct":           f"={base_no_K}*C{PM_COEF['debug_pct']}/100",
        "code_review_hours":   f"={base_no_K}*C{PM_COEF['code_review_hours']}/8",
        "communication_hours": f"={base_no_K}*C{PM_COEF['communication_hours']}/40",
        "qa_pct":              f"={base_no_K}*C{PM_COEF['qa_pct']}/100",
        "risk_buffer_pct":     f"={base_no_K}*C{PM_COEF['risk_buffer_pct']}/100",
        "devops_pct":          f"={base_no_K}*C{PM_COEF['devops_pct']}/100",
        "pm_pct":              f"=MAX(I{PM_SPEC_START}:I{PM_SPEC_END})*C{PM_COEF['pm_pct']}/100",
    }
    ws.cell(row=20, column=4, value="Недель вклад, инфо").font = FONT_HEADER
    ws.cell(row=20, column=4).alignment = ALIGN_CENTER
    for key, formula in d_formulas.items():
        r = PM_COEF[key]
        cell_d = ws.cell(row=r, column=4, value=formula)
        cell_d.font = FONT_NORMAL
        cell_d.alignment = ALIGN_RIGHT
        cell_d.border = BORDER_THIN
        cell_d.number_format = "0.00"

    # Grand total: project cost
    ws.merge_cells(start_row=20, start_column=7, end_row=20, end_column=10)
    hdr = ws.cell(row=20, column=7, value="💵 Себестоимость проекта")
    hdr.font = FONT_WHITE_BOLD
    hdr.fill = FILL_BLUE_HEADER
    hdr.alignment = ALIGN_CENTER
    hdr.border = BORDER_THIN

    ws.merge_cells(start_row=21, start_column=7, end_row=22, end_column=8)
    ws.merge_cells(start_row=21, start_column=9, end_row=22, end_column=10)
    label_cell = ws.cell(row=21, column=7, value="Итого, $")
    label_cell.font = FONT_BOLD
    label_cell.alignment = ALIGN_CENTER
    label_cell.fill = FILL_YELLOW

    total_cell = ws.cell(row=21, column=9,
                         value=f"=SUM(J{PM_QA_ROW}:J{PM_SPEC_END})")
    total_cell.font = FONT_TITLE
    total_cell.alignment = ALIGN_CENTER
    total_cell.fill = FILL_YELLOW
    total_cell.number_format = f"#,##0 \"{currency}\""

    _apply_border_rect(ws, 21, 7, 22, 10)


def _pm_write_auto_row(ws, row: int, name: str, rate, weeks_formula, cost_formula):
    """Write Manual QA / PM row on PM sheet."""
    ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
    ws.cell(row=row, column=2).alignment = ALIGN_LEFT
    ws.cell(row=row, column=2).border = BORDER_THIN
    # C: index — none for auto rows
    ws.cell(row=row, column=3).border = BORDER_THIN
    # D: rate editable
    d = ws.cell(row=row, column=4, value=rate)
    d.font = FONT_NORMAL
    d.fill = FILL_GREEN
    d.alignment = ALIGN_RIGHT
    d.border = BORDER_THIN
    d.number_format = "0.00"
    # E: mirror client rate from Sales
    sales_row = SALES_QA_ROW if row == PM_QA_ROW else SALES_MGR_ROW
    e = ws.cell(row=row, column=5, value=f"={SALES_REF}!K{sales_row}")
    e.font = FONT_NORMAL
    e.alignment = ALIGN_RIGHT
    e.border = BORDER_THIN
    e.number_format = "0.00"
    # G mirror
    g = ws.cell(row=row, column=7, value=f"=B{row}")
    g.font = FONT_NORMAL
    g.alignment = ALIGN_LEFT
    g.border = BORDER_THIN
    # H: empty (no numeric index for auto rows)
    ws.cell(row=row, column=8).border = BORDER_THIN
    # I: weeks formula
    i = ws.cell(row=row, column=9, value=weeks_formula)
    i.font = FONT_NORMAL
    i.alignment = ALIGN_RIGHT
    i.border = BORDER_THIN
    i.number_format = "0.00"
    # J: cost
    j = ws.cell(row=row, column=10, value=cost_formula)
    j.font = FONT_NORMAL
    j.alignment = ALIGN_RIGHT
    j.border = BORDER_THIN


# ═══════════════════════════════════════════════════════════════════════
# Sheet: Оценка (estimation) — formula-driven
# ═══════════════════════════════════════════════════════════════════════

def _build_estimation_sheet(wb, modules, roster, coefficients, task_to_est_row_out):
    ws = wb.create_sheet(ESTIMATION_SHEET)

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16

    # Instructions
    ws.cell(row=1, column=2, value="📊 Лист Оценка").font = FONT_TITLE
    ws.cell(row=2, column=2, value=(
        "Редактируйте только F/G (мин/макс дни по задачам). "
        "Колонка H и верхние итоги считаются формулами."
    )).font = FONT_NORMAL
    ws.row_dimensions[2].height = 20

    # Row 4: per-specialist summary headers (two blocks)
    hdr_style = lambda c: (setattr(c, 'font', FONT_WHITE_BOLD), c.fill == FILL_BLUE_HEADER,
                           setattr(c, 'fill', FILL_BLUE_HEADER),
                           setattr(c, 'alignment', ALIGN_CENTER),
                           setattr(c, 'border', BORDER_THIN))
    left_headers = [(1, "№"), (2, "Σ Мин дни"), (3, "Σ Макс дни"), (4, "Недель с K")]
    right_headers = [(5, "№"), (6, "Σ Мин дни"), (7, "Σ Макс дни"), (8, "Недель с K")]
    for col, val in left_headers + right_headers:
        c = ws.cell(row=4, column=col, value=val)
        c.font = FONT_WHITE_BOLD
        c.fill = FILL_BLUE_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
    ws.row_dimensions[4].height = 32

    # Rows 5..10: summary rows
    for i in range(6):
        r = EST_SUMMARY_START + i
        # Left block (indices 1..6)
        idx_left = i + 1
        left_spec = next((s for s in roster if s["index"] == idx_left), None)
        # A: index number
        ca = ws.cell(row=r, column=1, value=idx_left if left_spec else "")
        ca.font = FONT_NORMAL
        ca.alignment = ALIGN_CENTER
        ca.border = BORDER_THIN
        # B: SUMPRODUCT min
        cb = ws.cell(row=r, column=2, value=_sumproduct_formula(idx_left, "F"))
        cb.font = FONT_NORMAL
        cb.alignment = ALIGN_RIGHT
        cb.border = BORDER_THIN
        cb.number_format = "0.00"
        # C: SUMPRODUCT max
        cc = ws.cell(row=r, column=3, value=_sumproduct_formula(idx_left, "G"))
        cc.font = FONT_NORMAL
        cc.alignment = ALIGN_RIGHT
        cc.border = BORDER_THIN
        cc.number_format = "0.00"
        # D: weeks (mirror from PM!I{pm_row})
        if left_spec:
            cd = ws.cell(row=r, column=4, value=f"={PM_REF}!I{left_spec['pm_row']}")
        else:
            cd = ws.cell(row=r, column=4, value="")
        cd.font = FONT_BOLD
        cd.alignment = ALIGN_RIGHT
        cd.border = BORDER_THIN
        cd.number_format = "0.00"

        # Right block (indices 7..12)
        idx_right = i + 7
        right_spec = next((s for s in roster if s["index"] == idx_right), None)
        ce = ws.cell(row=r, column=5, value=idx_right if right_spec else "")
        ce.font = FONT_NORMAL
        ce.alignment = ALIGN_CENTER
        ce.border = BORDER_THIN
        cf = ws.cell(row=r, column=6, value=_sumproduct_formula(idx_right, "F"))
        cf.font = FONT_NORMAL
        cf.alignment = ALIGN_RIGHT
        cf.border = BORDER_THIN
        cf.number_format = "0.00"
        cg = ws.cell(row=r, column=7, value=_sumproduct_formula(idx_right, "G"))
        cg.font = FONT_NORMAL
        cg.alignment = ALIGN_RIGHT
        cg.border = BORDER_THIN
        cg.number_format = "0.00"
        if right_spec:
            ch = ws.cell(row=r, column=8, value=f"={PM_REF}!I{right_spec['pm_row']}")
        else:
            ch = ws.cell(row=r, column=8, value="")
        ch.font = FONT_BOLD
        ch.alignment = ALIGN_RIGHT
        ch.border = BORDER_THIN
        ch.number_format = "0.00"

    # Row 11: task table header
    task_hdr = {
        1: "№ спеца",
        2: "Вид работ",
        3: "Задача",
        4: "Комментарий",
        6: "Мин дни",
        7: "Макс дни",
        8: "Итого с K",
    }
    for col, val in task_hdr.items():
        c = ws.cell(row=EST_TASK_HEADER_ROW, column=col, value=val)
        c.font = FONT_WHITE_BOLD
        c.fill = FILL_BLUE_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
    ws.row_dimensions[EST_TASK_HEADER_ROW].height = 28

    # Task rows: group by module
    name_to_index = {s["name"]: s["index"] for s in roster}
    row = EST_TASK_START
    for module in modules:
        # Module header row (merged C:D)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        mc = ws.cell(row=row, column=3, value=module["name"])
        mc.font = FONT_SECTION
        mc.alignment = ALIGN_CENTER
        mc.fill = FILL_LIGHT_BLUE
        # Borders for the merged row
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = BORDER_THIN
            if col != 3 and col != 4:
                ws.cell(row=row, column=col).fill = FILL_LIGHT_BLUE
        row += 1

        tasks = module.get("tasks", [])
        for t_idx, t in enumerate(tasks):
            spec_name = t["specialist"]
            spec_index = name_to_index.get(spec_name, "")
            min_d = t.get("min_days", 0)
            max_d = t.get("max_days", 0)
            phase_tag = " [Post-MVP]" if t.get("phase") == "post-mvp" else ""

            # A: spec index
            ca = ws.cell(row=row, column=1, value=spec_index)
            ca.font = FONT_NORMAL
            ca.alignment = ALIGN_CENTER
            ca.border = BORDER_THIN
            # B: discipline name
            cb = ws.cell(row=row, column=2, value=spec_name)
            cb.font = FONT_BOLD
            cb.alignment = ALIGN_LEFT
            cb.border = BORDER_THIN
            # C: task
            cc = ws.cell(row=row, column=3, value=t["task"] + phase_tag)
            cc.font = FONT_NORMAL
            cc.alignment = ALIGN_LEFT
            cc.border = BORDER_THIN
            # D: comment
            cd = ws.cell(row=row, column=4, value=t.get("comment", ""))
            cd.font = FONT_NORMAL
            cd.alignment = ALIGN_LEFT
            cd.border = BORDER_THIN
            # F: min days (editable)
            cf = ws.cell(row=row, column=6, value=min_d)
            cf.font = FONT_NORMAL
            cf.fill = FILL_GREEN
            cf.alignment = ALIGN_CENTER
            cf.border = BORDER_THIN
            cf.number_format = "0.##"
            # G: max days (editable)
            cg = ws.cell(row=row, column=7, value=max_d)
            cg.font = FONT_NORMAL
            cg.fill = FILL_GREEN
            cg.alignment = ALIGN_CENTER
            cg.border = BORDER_THIN
            cg.number_format = "0.##"
            # H: formula = (F+G)/2 × K
            ch = ws.cell(row=row, column=8,
                         value=f"=IF(OR(F{row}=\"\",G{row}=\"\"),\"\",((F{row}+G{row})/2)*{k_expr()})")
            ch.font = FONT_NORMAL
            ch.alignment = ALIGN_CENTER
            ch.border = BORDER_THIN
            ch.number_format = "0.00"

            # Record mapping: which Оценка row holds this task (for Client/Bitrix formulas)
            task_to_est_row_out.append(row)
            row += 1


def _sumproduct_formula(index: int, col: str) -> str:
    """Sum col-values for tasks matching spec index.
    SUMIF is safer than SUMPRODUCT+IFERROR(range) — the latter schlumps the
    range to a scalar outside array-context in many Excel versions.
    """
    return (
        f"=SUMIF($A${EST_TASK_START}:$A${EST_TASK_END},{index},"
        f"${col}${EST_TASK_START}:${col}${EST_TASK_END})"
    )


# ═══════════════════════════════════════════════════════════════════════
# Sheet: Sales (Итоги оценки)
# ═══════════════════════════════════════════════════════════════════════

def _build_sales_sheet(wb, project_name, roster, rates, markup_pct, currency):
    ws = wb.create_sheet(SALES_SHEET)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 16
    ws.column_dimensions["M"].width = 12

    # Instructions
    ws.cell(row=1, column=2, value="💼 Лист Sales (Итоги оценки)").font = FONT_TITLE
    ws.cell(row=2, column=2, value=(
        "Зелёные ячейки редактируются: имя проекта, ссылки, маржа per-spec. "
        "Маржа = наценка в долях (1.0 = 100%, удвоение ставки клиенту)."
    )).font = FONT_NORMAL
    ws.row_dimensions[2].height = 24

    # Left block: project metadata
    ws.cell(row=4, column=2, value="Проект:").font = FONT_BOLD
    project_cell = ws.cell(row=4, column=3, value=project_name)
    project_cell.font = FONT_NORMAL
    project_cell.fill = FILL_GREEN
    project_cell.alignment = ALIGN_LEFT
    project_cell.border = BORDER_THIN

    ws.cell(row=5, column=2, value="Ссылка на макеты:").font = FONT_BOLD
    ws.cell(row=5, column=3).fill = FILL_GREEN
    ws.cell(row=5, column=3).border = BORDER_THIN

    ws.cell(row=6, column=2, value="Ссылка на ТЗ:").font = FONT_BOLD
    ws.cell(row=6, column=3).fill = FILL_GREEN
    ws.cell(row=6, column=3).border = BORDER_THIN

    ws.cell(row=9, column=2, value="Цена проекта клиенту:").font = FONT_BOLD
    price_cell = ws.cell(row=9, column=3, value="=J5")
    price_cell.font = FONT_TITLE
    price_cell.fill = FILL_YELLOW
    price_cell.alignment = ALIGN_RIGHT
    price_cell.border = BORDER_THIN
    price_cell.number_format = f"#,##0 \"{currency}\""

    ws.cell(row=10, column=2, value="Итоговая маржа %:").font = FONT_BOLD
    margin_cell = ws.cell(row=10, column=3, value="=IFERROR((J5-I5)/J5,0)")
    margin_cell.font = FONT_BOLD
    margin_cell.alignment = ALIGN_RIGHT
    margin_cell.border = BORDER_THIN
    margin_cell.number_format = "0.0%"

    ws.cell(row=12, column=2, value="Часы всего:").font = FONT_BOLD
    hours_cell = ws.cell(row=12, column=3, value="=H5")
    hours_cell.font = FONT_NORMAL
    hours_cell.alignment = ALIGN_RIGHT
    hours_cell.border = BORDER_THIN

    ws.cell(row=13, column=2, value="Дней всего:").font = FONT_BOLD
    days_cell = ws.cell(row=13, column=3, value="=G5")
    days_cell.font = FONT_NORMAL
    days_cell.alignment = ALIGN_RIGHT
    days_cell.border = BORDER_THIN

    ws.cell(row=14, column=2, value="Недель всего:").font = FONT_BOLD
    weeks_cell = ws.cell(row=14, column=3, value="=F5")
    weeks_cell.font = FONT_NORMAL
    weeks_cell.alignment = ALIGN_RIGHT
    weeks_cell.border = BORDER_THIN

    # Right block headers (row 4)
    right_hdr = {
        5: "Проектная команда",
        6: "Недель",
        7: "Дней",
        8: "Часов",
        9: f"Себестоимость, {currency}",
        10: f"Стоимость работ, {currency}",
        11: f"Ставка клиенту, {currency}",
        12: f"Ставка на руки, {currency}",
        13: "Маржа (доля)",
    }
    for col, val in right_hdr.items():
        c = ws.cell(row=4, column=col, value=val)
        c.font = FONT_WHITE_BOLD
        c.fill = FILL_BLUE_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    # Row 5: totals row
    ws.cell(row=SALES_TOTAL_ROW, column=5, value="ИТОГО по проекту:").font = FONT_BOLD
    ws.cell(row=SALES_TOTAL_ROW, column=5).fill = FILL_YELLOW
    ws.cell(row=SALES_TOTAL_ROW, column=5).alignment = ALIGN_RIGHT
    ws.cell(row=SALES_TOTAL_ROW, column=5).border = BORDER_THIN
    ws.cell(row=SALES_TOTAL_ROW, column=6, value=f"=SUM(F{SALES_QA_ROW}:F{SALES_SPEC_END})").font = FONT_BOLD
    ws.cell(row=SALES_TOTAL_ROW, column=7, value=f"=F{SALES_TOTAL_ROW}*5").font = FONT_BOLD
    ws.cell(row=SALES_TOTAL_ROW, column=8, value=f"=G{SALES_TOTAL_ROW}*8").font = FONT_BOLD
    ws.cell(row=SALES_TOTAL_ROW, column=9, value=f"=SUM(I{SALES_QA_ROW}:I{SALES_SPEC_END})").font = FONT_BOLD
    ws.cell(row=SALES_TOTAL_ROW, column=10, value=f"=SUM(J{SALES_QA_ROW}:J{SALES_SPEC_END})").font = FONT_BOLD
    for col in range(6, 11):
        c = ws.cell(row=SALES_TOTAL_ROW, column=col)
        c.fill = FILL_YELLOW
        c.alignment = ALIGN_RIGHT
        c.border = BORDER_THIN
        c.number_format = "0.00" if col < 9 else f"#,##0 \"{currency}\""
    for col in (11, 12, 13):
        ws.cell(row=SALES_TOTAL_ROW, column=col).fill = FILL_YELLOW
        ws.cell(row=SALES_TOTAL_ROW, column=col).border = BORDER_THIN

    # Default markup as ratio (1.0 = 100%)
    markup_ratio = markup_pct / 100.0

    # Rows: Manual QA, PM, then real specialists
    _sales_write_aggregate_row(
        ws, SALES_QA_ROW,
        name_formula=f"={PM_REF}!B{PM_QA_ROW}",
        pm_row=PM_QA_ROW, currency=currency, markup_ratio=markup_ratio,
    )
    _sales_write_aggregate_row(
        ws, SALES_MGR_ROW,
        name_formula=f"={PM_REF}!B{PM_MGR_ROW}",
        pm_row=PM_MGR_ROW, currency=currency, markup_ratio=markup_ratio,
    )
    for s in roster:
        sales_row = SALES_SPEC_START + s["index"] - 1
        _sales_write_aggregate_row(
            ws, sales_row,
            name_formula=f"={PM_REF}!B{s['pm_row']}",
            pm_row=s["pm_row"], currency=currency, markup_ratio=markup_ratio,
        )
    # Empty remaining slots
    for i in range(len(roster), PM_SPEC_MAX):
        r = SALES_SPEC_START + i
        for col in range(5, 14):
            ws.cell(row=r, column=col).border = BORDER_THIN
        ws.cell(row=r, column=13).fill = FILL_GREEN

    ws.row_dimensions[4].height = 32


def _sales_write_aggregate_row(ws, row, name_formula, pm_row, currency, markup_ratio):
    e = ws.cell(row=row, column=5, value=name_formula)
    e.font = FONT_NORMAL
    e.alignment = ALIGN_LEFT
    e.border = BORDER_THIN

    f = ws.cell(row=row, column=6, value=f"={PM_REF}!I{pm_row}")
    f.font = FONT_NORMAL
    f.alignment = ALIGN_RIGHT
    f.border = BORDER_THIN
    f.number_format = "0.00"

    g = ws.cell(row=row, column=7, value=f"=F{row}*5")
    g.font = FONT_NORMAL
    g.alignment = ALIGN_RIGHT
    g.border = BORDER_THIN
    g.number_format = "0.0"

    h = ws.cell(row=row, column=8, value=f"=G{row}*8")
    h.font = FONT_NORMAL
    h.alignment = ALIGN_RIGHT
    h.border = BORDER_THIN
    h.number_format = "0"

    i = ws.cell(row=row, column=9, value=f"={PM_REF}!J{pm_row}")
    i.font = FONT_NORMAL
    i.alignment = ALIGN_RIGHT
    i.border = BORDER_THIN
    i.number_format = f"#,##0 \"{currency}\""

    j = ws.cell(row=row, column=10, value=f"=I{row}*(1+M{row})")
    j.font = FONT_BOLD
    j.alignment = ALIGN_RIGHT
    j.border = BORDER_THIN
    j.number_format = f"#,##0 \"{currency}\""

    k = ws.cell(row=row, column=11, value=f"={PM_REF}!D{pm_row}*(1+M{row})")
    k.font = FONT_NORMAL
    k.alignment = ALIGN_RIGHT
    k.border = BORDER_THIN
    k.number_format = "0.00"

    l = ws.cell(row=row, column=12, value=f"={PM_REF}!D{pm_row}")
    l.font = FONT_NORMAL
    l.alignment = ALIGN_RIGHT
    l.border = BORDER_THIN
    l.number_format = "0.00"

    m = ws.cell(row=row, column=13, value=markup_ratio)
    m.font = FONT_NORMAL
    m.fill = FILL_GREEN
    m.alignment = ALIGN_CENTER
    m.border = BORDER_THIN
    m.number_format = "0%"


# ═══════════════════════════════════════════════════════════════════════
# Sheet: Для клиента — formulas referencing Sales/Оценка
# ═══════════════════════════════════════════════════════════════════════

def _build_client_sheet(wb, project_name, modules, roster, rates, markup_pct, currency, task_to_est_row_placeholder):
    """Build client-facing sheet. task_to_est_row_placeholder is populated
    later by Оценка builder; we post-fill task rows via a second pass.

    To keep the flow simple, we compute est_row inline using the same
    traversal order that Оценка will use.
    """
    ws = wb.active
    ws.title = CLIENT_SHEET

    ws.column_dimensions["A"].width = 7.8
    ws.column_dimensions["B"].width = 21
    ws.column_dimensions["C"].width = 39.5
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 20

    font_title = Font(name="Montserrat", size=24)
    font_subtitle = Font(name="Montserrat", size=10)
    font_date = Font(name="Montserrat", size=10, color="777777")
    font_header = Font(name="Montserrat", size=12)
    font_normal = Font(name="Montserrat", size=10)
    font_section = Font(name="Montserrat", size=11, bold=True)

    # Logo
    logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "metalamp-logo.png")
    if os.path.exists(logo_path):
        from openpyxl.drawing.image import Image as XlImage
        img = XlImage(logo_path)
        orig_w, orig_h = img.width, img.height
        target_w = 300
        img.width = target_w
        img.height = int(orig_h * (target_w / max(orig_w, 1)))
        ws.add_image(img, "B2")

    # Title: row 5 reads project name from Sales
    ws.row_dimensions[5].height = 38.25
    ws.merge_cells("B5:C5")
    ws.cell(row=5, column=2, value="Оценка проекта ").font = font_title
    ws.merge_cells("D5:E5")
    ws.cell(row=5, column=4, value=f"={SALES_REF}!C4").font = font_title

    ws.row_dimensions[6].height = 38.25
    ws.merge_cells("B6:E6")
    ws.cell(row=6, column=2, value="В стоимость входит тестирование, работа менеджера").font = font_subtitle

    ws.row_dimensions[7].height = 31.5
    ws.merge_cells("B7:E7")
    ws.cell(row=7, column=2,
            value="В течение 3 месяцев мы бесплатно устраняем технические ошибки (техподдержка)").font = font_subtitle

    ws.cell(row=8, column=2,
            value=f"Актуально на: {date.today().strftime('%d.%m.%Y')}").font = font_date

    # Summary block (B9:D10) — all three values via formulas
    ws.row_dimensions[9].height = 30
    for col, val in [(2, "Команда проекта,\nчеловек"), (3, "Длительность проекта,\nчасы"),
                     (4, f"Стоимость, {currency}")]:
        c = ws.cell(row=9, column=col, value=val)
        c.font = font_header
        c.fill = FILL_ORANGE
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    # Team count = number of Sales rows with positive weeks (Manual QA..specs)
    ws.cell(row=10, column=2,
            value=f"=SUMPRODUCT(--({SALES_REF}!F{SALES_QA_ROW}:F{SALES_SPEC_END}>0))")
    ws.cell(row=10, column=3, value=f"={SALES_REF}!H{SALES_TOTAL_ROW}")
    ws.cell(row=10, column=3).number_format = "0"
    ws.cell(row=10, column=4, value=f"={SALES_REF}!J{SALES_TOTAL_ROW}")
    ws.cell(row=10, column=4).number_format = f"#,##0"
    for col in (2, 3, 4):
        c = ws.cell(row=10, column=col)
        c.font = font_normal
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    # Specialists table (row 12+)
    ws.row_dimensions[12].height = 15
    for col, val in [(2, "Специалисты"), (3, "Занятость, недели"),
                     (4, "Занятость, часы"), (5, f"Стоимость, {currency}")]:
        c = ws.cell(row=12, column=col, value=val)
        c.font = font_header
        c.fill = FILL_ORANGE
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN

    # Rows: Manual QA, PM, real specs — reading from Sales
    sales_rows_to_show = [SALES_QA_ROW, SALES_MGR_ROW] + [
        SALES_SPEC_START + s["index"] - 1 for s in roster
    ]
    client_row = 13
    for idx, sales_row in enumerate(sales_rows_to_show):
        fill = FILL_GRAY if idx % 2 == 0 else FILL_WHITE
        vals = [
            (2, f"={SALES_REF}!E{sales_row}"),
            (3, f"={SALES_REF}!F{sales_row}"),
            (4, f"={SALES_REF}!H{sales_row}"),
            (5, f"={SALES_REF}!J{sales_row}"),
        ]
        for col, formula in vals:
            c = ws.cell(row=client_row, column=col, value=formula)
            c.font = font_normal
            c.fill = fill
            c.alignment = ALIGN_CENTER if col >= 3 else ALIGN_LEFT
            c.border = BORDER_THIN
            if col == 3:
                c.number_format = "0.0"
            elif col == 4:
                c.number_format = "0"
            elif col == 5:
                c.number_format = f"#,##0 \"{currency}\""
        client_row += 1

    # Task decomposition (per module, one sub-table each)
    # Mirrors Оценка traversal to compute the right row number
    row = client_row + 2
    est_row_counter = EST_TASK_START   # matches Оценка builder
    for module in modules:
        ws.row_dimensions[row].height = 13.8
        ws.cell(row=row, column=2, value=module["name"]).font = font_section
        row += 1

        # Module occupies one row in Оценка (merged C:D); skip that counter
        est_row_counter += 1

        ws.row_dimensions[row].height = 15
        for col, val in [(2, "Специалист"), (3, "Задача"),
                         (4, "Комментарий"), (5, "Оценка, дни")]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = font_header
            c.fill = FILL_ORANGE
            c.alignment = ALIGN_CENTER
            c.border = BORDER_THIN
        row += 1

        for t_idx, t in enumerate(module.get("tasks", [])):
            fill = FILL_GRAY if t_idx % 2 == 0 else FILL_WHITE
            phase_tag = " [Post-MVP]" if t.get("phase") == "post-mvp" else ""
            # Columns B-D — static from decomposition
            ws.cell(row=row, column=2, value=t["specialist"]).font = font_normal
            ws.cell(row=row, column=2).fill = fill
            ws.cell(row=row, column=2).alignment = ALIGN_LEFT
            ws.cell(row=row, column=2).border = BORDER_THIN

            ws.cell(row=row, column=3, value=t["task"] + phase_tag).font = font_normal
            ws.cell(row=row, column=3).fill = fill
            ws.cell(row=row, column=3).alignment = ALIGN_LEFT
            ws.cell(row=row, column=3).border = BORDER_THIN

            ws.cell(row=row, column=4, value=t.get("comment", "")).font = font_normal
            ws.cell(row=row, column=4).fill = fill
            ws.cell(row=row, column=4).alignment = ALIGN_LEFT
            ws.cell(row=row, column=4).border = BORDER_THIN

            # E: integer days via ROUND formula from Оценка H (matches Gantt half-up round)
            days_formula = (
                f'=IF(ISBLANK({EST_REF}!H{est_row_counter}),"",'
                f'MAX(1,ROUND({EST_REF}!H{est_row_counter},0)))'
            )
            c_days = ws.cell(row=row, column=5, value=days_formula)
            c_days.font = font_normal
            c_days.fill = fill
            c_days.alignment = ALIGN_CENTER
            c_days.border = BORDER_THIN
            c_days.number_format = "0"

            row += 1
            est_row_counter += 1
        row += 1


# ═══════════════════════════════════════════════════════════════════════
# Sheet: Для Битрикса — flat CRM-import format
# ═══════════════════════════════════════════════════════════════════════

def _build_bitrix_sheet(wb, modules, task_to_est_row_placeholder):
    ws = wb.create_sheet(BITRIX_SHEET)

    # Column setup — match Bitrix24 task import template exactly
    headers = [
        "Название",
        "Описание",
        "Это важная задача",
        "Разрешить ответственному менять сроки задачи",
        "Пропускать выходные и праздничные дни",
        "Проконтролировать задачу после завершения",
        "Проект",
        "Задано время на завершение задачи",
        "Время на выполнение задачи в секундах",
        "Теги",
        "Начать задачу с",
        "Завершить задачу",
    ]
    widths = [45, 60, 8, 8, 8, 8, 30, 10, 12, 20, 14, 14]
    for i, (h, w) in enumerate(zip(headers, widths)):
        col = i + 1
        ws.column_dimensions[get_column_letter(col)].width = w
        c = ws.cell(row=1, column=col, value=h)
        c.font = FONT_WHITE_BOLD
        c.fill = FILL_BLUE_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_THIN
    ws.row_dimensions[1].height = 32

    # Flatten all tasks; for each, reference the Оценка task row.
    # Module-header rows in Оценка don't produce Bitrix rows.
    row = 2
    est_row = EST_TASK_START
    for module in modules:
        est_row += 1   # module header row in Оценка
        for t in module.get("tasks", []):
            # A: Название (from Оценка C)
            a = ws.cell(row=row, column=1, value=f"={EST_REF}!C{est_row}")
            a.font = FONT_NORMAL
            a.alignment = ALIGN_LEFT
            a.border = BORDER_THIN
            # B: Описание (from Оценка D)
            b = ws.cell(row=row, column=2, value=f"={EST_REF}!D{est_row}")
            b.font = FONT_NORMAL
            b.alignment = ALIGN_LEFT
            b.border = BORDER_THIN
            # C-F and H: boolean-style flags (0 or 1) only if A populated
            for col, flag_value in [(3, 0), (4, 0), (5, 1), (6, 1), (8, 1)]:
                c = ws.cell(row=row, column=col,
                            value=f'=IF(A{row}<>"",{flag_value},"")')
                c.font = FONT_NORMAL
                c.alignment = ALIGN_CENTER
                c.border = BORDER_THIN
            # G: Project name (from Sales C4)
            g = ws.cell(row=row, column=7,
                        value=f'=IF(A{row}<>"",{SALES_REF}!$C$4,"")')
            g.font = FONT_NORMAL
            g.alignment = ALIGN_LEFT
            g.border = BORDER_THIN
            # I: duration in seconds — Оценка H (days) × 8 × 3600
            i = ws.cell(row=row, column=9,
                        value=f'=IF(ISBLANK({EST_REF}!H{est_row}),"",{EST_REF}!H{est_row}*8*3600)')
            i.font = FONT_NORMAL
            i.alignment = ALIGN_RIGHT
            i.border = BORDER_THIN
            i.number_format = "0"
            # J: tags = discipline (Оценка B)
            j = ws.cell(row=row, column=10, value=f"={EST_REF}!B{est_row}")
            j.font = FONT_NORMAL
            j.alignment = ALIGN_LEFT
            j.border = BORDER_THIN
            # K: start = today
            k = ws.cell(row=row, column=11, value=f'=IF(A{row}<>"",TODAY(),"")')
            k.font = FONT_NORMAL
            k.alignment = ALIGN_CENTER
            k.border = BORDER_THIN
            k.number_format = "DD.MM.YYYY"
            # L: end = today + integer workdays from Оценка H (rounded same way as client sheet)
            l = ws.cell(row=row, column=12,
                        value=(f'=IF(ISBLANK({EST_REF}!H{est_row}),"",'
                               f'WORKDAY(TODAY(),MAX(1,ROUND({EST_REF}!H{est_row},0))))'))
            l.font = FONT_NORMAL
            l.alignment = ALIGN_CENTER
            l.border = BORDER_THIN
            l.number_format = "DD.MM.YYYY"

            row += 1
            est_row += 1


# ═══════════════════════════════════════════════════════════════════════
# Sheet: GANTT Chart — integer days with half-up rounding
# ═══════════════════════════════════════════════════════════════════════

def _build_gantt_sheet(wb, modules, K, roster=None):
    def next_workday(dt):
        while dt.weekday() >= 5:
            dt += timedelta(days=1)
        return dt

    def add_workdays(start_dt, num_workdays):
        cur = next_workday(start_dt)
        counted = 1
        while counted < num_workdays:
            cur += timedelta(days=1)
            if cur.weekday() < 5:
                counted += 1
        return cur

    PHASE_COLORS = [
        {"header": "1F4E79", "fill": "D6E4F0", "bar": "5B9BD5"},
        {"header": "7B2D26", "fill": "F2DCDB", "bar": "C0504D"},
        {"header": "4F6228", "fill": "EBF1DE", "bar": "9BBB59"},
        {"header": "31859C", "fill": "DAEEF3", "bar": "4BACC6"},
        {"header": "E36C09", "fill": "FDE9D9", "bar": "F79646"},
        {"header": "60497A", "fill": "E4DFEC", "bar": "8064A2"},
        {"header": "4A452A", "fill": "F2F2E6", "bar": "948A54"},
    ]

    month_names_ru = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
    }

    DATA_COL_START = 7

    specialist_tasks: dict[str, list[dict]] = {}
    for module in modules:
        for t in module.get("tasks", []):
            spec = t["specialist"]
            specialist_tasks.setdefault(spec, [])
            avg_d = (t.get("min_days", 0) + t.get("max_days", 0)) / 2
            duration_days = max(1, half_up(avg_d * K))   # ⬅ half-away-from-zero (Excel-compatible)
            specialist_tasks[spec].append({"task": t["task"], "duration": duration_days})

    if not specialist_tasks:
        return

    spec_list = list(specialist_tasks.keys())
    spec_colors = {s: PHASE_COLORS[i % len(PHASE_COLORS)] for i, s in enumerate(spec_list)}

    project_start_raw = next_workday(date.today())
    project_start_dt = datetime(project_start_raw.year, project_start_raw.month, project_start_raw.day)

    scheduled_tasks = []
    for spec in spec_list:
        current_start = project_start_dt
        for t in specialist_tasks[spec]:
            start = next_workday(current_start)
            end = add_workdays(start, t["duration"])
            scheduled_tasks.append((spec, t["task"], start, end, t["duration"]))
            current_start = end + timedelta(days=1)

    if not scheduled_tasks:
        return

    project_start = min(t[2] for t in scheduled_tasks)
    while project_start.weekday() != 0:
        project_start -= timedelta(days=1)
    project_end = max(t[3] for t in scheduled_tasks)
    while project_end.weekday() != 4:
        project_end += timedelta(days=1)

    all_days = []
    d = project_start
    while d <= project_end:
        all_days.append(d)
        d += timedelta(days=1)
    num_days = len(all_days)

    ws = wb.create_sheet(GANTT_SHEET)

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    weekend_header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    date_label_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    thin_border = Border(left=Side("thin", color="BFBFBF"), right=Side("thin", color="BFBFBF"),
                         top=Side("thin", color="BFBFBF"), bottom=Side("thin", color="BFBFBF"))
    week_sep_border = Border(left=Side("thin", color="BFBFBF"), right=Side("medium", color="808080"),
                             top=Side("thin", color="BFBFBF"), bottom=Side("thin", color="BFBFBF"))

    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    month_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    day_num_font = Font(name="Calibri", size=7, color="44546A")
    weekday_font = Font(name="Calibri", size=6, color="808080")
    weekday_bold_font = Font(name="Calibri", size=6, bold=True, color="999999")
    task_font = Font(name="Calibri", size=9)
    phase_header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    center_align = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(vertical="center", wrap_text=True)
    phase_align = Alignment(horizontal="left", vertical="center")

    weekday_names = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 11
    ws.column_dimensions["E"].width = 6
    ws.column_dimensions["F"].width = 11
    for i in range(num_days):
        col_letter = get_column_letter(DATA_COL_START + i)
        ws.column_dimensions[col_letter].width = 2.5 if all_days[i].weekday() >= 5 else 3.8

    def _border_for_day(day):
        return week_sep_border if day.weekday() == 6 else thin_border

    ws.row_dimensions[1].height = 20
    for c in range(1, 7):
        ws.cell(row=1, column=c).fill = header_fill
        ws.cell(row=1, column=c).border = thin_border

    month_spans = []
    if all_days:
        cur_month = (all_days[0].year, all_days[0].month)
        span_start = 0
        for i, day in enumerate(all_days):
            m = (day.year, day.month)
            if m != cur_month:
                month_spans.append((cur_month, span_start, i - 1))
                cur_month = m
                span_start = i
        month_spans.append((cur_month, span_start, len(all_days) - 1))

    for (year, month), start_idx, end_idx in month_spans:
        start_col = DATA_COL_START + start_idx
        end_col = DATA_COL_START + end_idx
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(row=1, column=start_col, value=f"{month_names_ru[month]} {year}")
        cell.font = month_font
        cell.fill = header_fill
        cell.alignment = center_align
        for ci in range(start_col, end_col + 1):
            c = ws.cell(row=1, column=ci)
            c.fill = header_fill
            c.border = _border_for_day(all_days[ci - DATA_COL_START])

    ws.row_dimensions[2].height = 18
    for i, val in enumerate(["Фаза", "Роль", "Задача", "Старт", "Дней", "Конец"]):
        cell = ws.cell(row=2, column=i + 1, value=val)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, day in enumerate(all_days):
        col = DATA_COL_START + i
        cell = ws.cell(row=2, column=col, value=day.day)
        cell.font = day_num_font
        cell.fill = weekend_header_fill if day.weekday() >= 5 else date_label_fill
        cell.alignment = center_align
        cell.border = _border_for_day(day)

    ws.row_dimensions[3].height = 14
    for c in range(1, 7):
        ws.cell(row=3, column=c).fill = date_label_fill
        ws.cell(row=3, column=c).border = thin_border
    for i, day in enumerate(all_days):
        col = DATA_COL_START + i
        cell = ws.cell(row=3, column=col, value=weekday_names[day.weekday()])
        cell.font = weekday_bold_font if day.weekday() >= 5 else weekday_font
        cell.fill = weekend_header_fill if day.weekday() >= 5 else date_label_fill
        cell.alignment = center_align
        cell.border = _border_for_day(day)

    row = 4
    current_spec = None
    for spec, task_name, start_dt, end_dt, duration in scheduled_tasks:
        colors = spec_colors[spec]
        rb = int(colors["fill"][:2], 16)
        gb = int(colors["fill"][2:4], 16)
        bb = int(colors["fill"][4:6], 16)
        wknd_hex = f"{max(0, rb - 25):02X}{max(0, gb - 25):02X}{max(0, bb - 25):02X}"

        fill_phase = PatternFill(start_color=colors["fill"], end_color=colors["fill"], fill_type="solid")
        fill_bar = PatternFill(start_color=colors["bar"], end_color=colors["bar"], fill_type="solid")
        fill_wknd = PatternFill(start_color=wknd_hex, end_color=wknd_hex, fill_type="solid")
        fill_header = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")

        if spec != current_spec:
            current_spec = spec
            ws.row_dimensions[row].height = 22
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            cell = ws.cell(row=row, column=1, value=spec)
            cell.font = phase_header_font
            cell.fill = fill_header
            cell.alignment = phase_align
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = fill_header
                ws.cell(row=row, column=c).border = thin_border
            for i, day in enumerate(all_days):
                col = DATA_COL_START + i
                ws.cell(row=row, column=col).fill = fill_header
                ws.cell(row=row, column=col).border = _border_for_day(day)
            row += 1

        ws.row_dimensions[row].height = 20
        task_data = [
            (1, spec, left_wrap),
            (2, spec, left_wrap),
            (3, task_name, left_wrap),
            (4, start_dt.strftime("%d.%m.%y"), center_align),
            (5, duration, center_align),
            (6, end_dt.strftime("%d.%m.%y"), center_align),
        ]
        for col, val, align in task_data:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = task_font
            cell.fill = fill_phase
            cell.alignment = align
            cell.border = thin_border

        for i, day in enumerate(all_days):
            col = DATA_COL_START + i
            cell = ws.cell(row=row, column=col)
            cell.border = _border_for_day(day)
            if start_dt <= day <= end_dt and day.weekday() < 5:
                cell.fill = fill_bar
            elif day.weekday() >= 5:
                cell.fill = fill_wknd
            else:
                cell.fill = fill_phase
        row += 1

    ws.freeze_panes = "G4"


# ═══════════════════════════════════════════════════════════════════════
# SIMPLE MODE (fallback: 3 sheets, no formulas, days only)
# ═══════════════════════════════════════════════════════════════════════

def _build_simple_mode(wb, project_name, modules, K):
    """Legacy path: Для клиента + Оценка + GANTT, no formulas, no money."""
    specs = _simple_compute_specialists(modules, K)

    ws = wb.active
    ws.title = CLIENT_SHEET
    _simple_client_sheet(ws, project_name, modules, K, specs)
    _simple_estimation_sheet(wb, modules, K, specs)
    _build_gantt_sheet(wb, modules, K)


def _simple_compute_specialists(modules, K):
    spec_days = {}
    for m in modules:
        for t in m.get("tasks", []):
            name = t["specialist"]
            avg = (t.get("min_days", 0) + t.get("max_days", 0)) / 2
            spec_days[name] = spec_days.get(name, 0) + avg
    result = []
    for name, days in spec_days.items():
        adj = days * K
        result.append({
            "name": name,
            "days": round(days, 1),
            "adjusted_days": round(adj, 1),
            "hours": round(adj * 8, 1),
            "weeks": round(adj / 5, 1),
        })
    return result


def _simple_client_sheet(ws, project_name, modules, K, specs):
    ws.column_dimensions["A"].width = 7.8
    ws.column_dimensions["B"].width = 21
    ws.column_dimensions["C"].width = 39.5
    ws.column_dimensions["D"].width = 35

    ws.row_dimensions[5].height = 38.25
    ws.merge_cells("B5:C5")
    ws.cell(row=5, column=2, value="Оценка проекта ").font = FONT_TITLE
    ws.merge_cells("D5:E5")
    ws.cell(row=5, column=4, value=project_name).font = FONT_TITLE

    ws.cell(row=8, column=2, value=f"Актуально на: {date.today().strftime('%d.%m.%Y')}")

    total_hours = sum(s["hours"] for s in specs)
    team_size = len(specs)

    for col, val in [(2, "Команда"), (3, "Часы")]:
        c = ws.cell(row=9, column=col, value=val)
        c.font = FONT_HEADER
        c.fill = FILL_ORANGE
        c.alignment = ALIGN_CENTER
    ws.cell(row=10, column=2, value=team_size)
    ws.cell(row=10, column=3, value=round(total_hours))

    row = 12
    for col, val in [(2, "Специалисты"), (3, "Недели"), (4, "Часы")]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = FONT_HEADER
        c.fill = FILL_ORANGE
        c.alignment = ALIGN_CENTER
    for idx, s in enumerate(specs):
        r = row + 1 + idx
        ws.cell(row=r, column=2, value=s["name"])
        ws.cell(row=r, column=3, value=s["weeks"])
        ws.cell(row=r, column=4, value=round(s["hours"]))


def _simple_estimation_sheet(wb, modules, K, specs):
    ws = wb.create_sheet(ESTIMATION_SHEET)
    ws.cell(row=1, column=1, value="Простой режим (без формул)").font = FONT_HEADER
    row = 3
    for m in modules:
        ws.cell(row=row, column=1, value=m["name"]).font = FONT_BOLD
        row += 1
        for t in m.get("tasks", []):
            avg = (t.get("min_days", 0) + t.get("max_days", 0)) / 2
            ws.cell(row=row, column=1, value=t["specialist"])
            ws.cell(row=row, column=2, value=t["task"])
            ws.cell(row=row, column=3, value=round(avg * K, 1))
            row += 1


# ═══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python scripts/build_xlsx.py <decomposition.json> <output.xlsx> "
              "[--params <estimation_params.json>] [--K 1.0] [--name \"Project Name\"]")
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)

    params = None
    if "--params" in sys.argv:
        with open(sys.argv[sys.argv.index("--params") + 1], "r", encoding="utf-8") as f:
            params = json.load(f)

    K = 1.0
    if "--K" in sys.argv:
        K = float(sys.argv[sys.argv.index("--K") + 1])

    modules = data.get("modules", data) if isinstance(data, dict) else data

    project_name = "Проект"
    if "--name" in sys.argv:
        project_name = sys.argv[sys.argv.index("--name") + 1]
    elif isinstance(data, dict) and "project_name" in data:
        project_name = data["project_name"]

    result = build_report_xlsx(project_name, modules, K, params)

    with open(sys.argv[2], "wb") as f:
        f.write(result.read())

    mode = "full (6 sheets, formulas)" if params else "simple (3 sheets, days only)"
    print(f"Saved: {sys.argv[2]} [{mode}]")
